import calendar
import datetime
import argparse
import os

from collections import defaultdict

import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook

import dretail
from branch import Branch

parser = argparse.ArgumentParser()
parser.add_argument('year', nargs='?', type=int, default=datetime.datetime.today().year)
parser.add_argument('month', nargs='?', type=int, default=datetime.datetime.today().month)
parser.add_argument('day', nargs='?', type=int, default=datetime.datetime.today().day)
args = parser.parse_args()

branch = Branch(
    ["ODczMw==", "MTMxODI=", ],
    [16284, 16179, ],
    'SERANG',
    ['582E227AX2XLTOJCDDGAQX84O1CA8C39', '865038d2211cd111b16dd23e39ea342f', ],
    ['5b0c7b868ca031f555e490e9b29fd8de', '6549F75F84PCI8HOCAIWBM2P5E3K070F', ],
    ['a%3A4%3A%7Bs%3A10%3A%22session_id%22%3Bs%3A32%3A%22e1f12245cd75ec9cb853eb10efe2cc47%22%3Bs%3A10%3A%22ip_address%22%3Bs%3A13%3A%2210.100.100.62%22%3Bs%3A10%3A%22user_agent%22%3Bs%3A111%3A%22Mozilla%2F5.0+%28Windows+NT+10.0%3B+Win64%3B+x64%29+AppleWebKit%2F537.36+%28KHTML%2C+like+Gecko%29+Chrome%2F111.0.0.0+Safari%2F537.36%22%3Bs%3A13%3A%22last_activity%22%3Bi%3A1681234190%3B%7Dd6f22faf3f2fdf22bcb4c85ccbe5e18db99b45b5',
     'b8YFauEf1FfJjVu%2Fx1wHRvYBDnbRS6CKIQBAYUDqyWK0nszbtZMb2HuUILEBZsNHo%2FRZRGOjhaH4LGkj%2BlQlu5Uy55iSOsJOydM45NJlbf2X6RfQG%2Fj2fLc7LeHYlxkp2C0S9A%2F5Rrs7O5%2BxQXWdX41Qmgir6FLKSu%2Bkt4aoVShlRgXAesQHjDPRS3sHSRr3CCQCRz%2BFTyWG1fUTObOMVdcrfmE9Qmfz6Qs6SGOoxsHnOHDenAhDFdY3UdTaCgEsJZ9xbH5vrQR0F3%2FSfYVsm6QkZEQkCuZqaNE8G1tCMC8RDqZGwOK54635aB5jx8ViSH4USbax2cpY9RJ0laC1bKa321gp7Kr%2BE9RMTqp7sw3AJfIhxFzL0evIgmVriCmonqkGN16rJv2yE1aBNDrlvrHjOlDutEEiZMmwx7JVFbM%3Dbfd16832cae10478f51a7e7257ac9c15087e5a93', ],
    0,
    '',
    10,
    3,
    (1000, 1000, 1000, 1000, 1500, 2000, 1500,),
)

modifiers = {
    'Hot Chocolate': {
        'Mint': 'Hot Chocolate',
        'Strawberry': 'Hot Chocolate',
    },
    'Iced Chocolate': {
        'Mint': 'Iced Chocolate Mint',
        'Strawberry': 'Iced Chocolate Berry',
    },
    'Iced Chocolate Large': {
        'Mint': 'Iced Chocolate Mint Large',
        'Strawberry': 'Iced Chocolate Berry Large',
    },
    'Pisang': {
        'Keju': 'Pisang Keju',
    },
}

filename = f'Sales Data {args.year}.xlsx'
if os.path.exists(filename):
    wb = openpyxl.load_workbook(filename)
else:
    wb = Workbook()
    wb.save(filename)

title = f'{calendar.month_name[args.month]} {args.year}'
if title in wb.sheetnames:
    ws = wb[title]
else:
    ws = wb.create_sheet(title)

start_time = datetime.datetime(args.year, args.month, args.day, 10, 0, 0)
end_time = start_time + relativedelta(days=1, hour=6)
current_time = start_time

while current_time < end_time:
    shifting_start_time = current_time
    shifting_end_time = current_time + relativedelta(days=1, hour=6)
    cell_row = 1
    ws.cell(row=1, column=shifting_start_time.day + 1, value=shifting_start_time.day)
    print(shifting_start_time)

    while current_time < shifting_end_time:
        menus = defaultdict(int)
        current_end_time = current_time + relativedelta(minute=59, second=59)

        for backoffice in branch.backoffices:
            df_sales_report = dretail.get_laporan_sales_by_item_detail(
                backoffice.branch_id,
                backoffice.company_id,
                current_time.strftime('%d/%m/%Y'),
                current_time.strftime('%H:%M'),
                current_end_time.strftime('%d/%m/%Y'),
                current_end_time.strftime('%H:%M'),
                backoffice.get_cookies(),
            )

            for _, row in df_sales_report.iterrows():
                modifier = str(row['Modifier']).rstrip(' 1X')
                item_name = row['Item Name'].rstrip('.')

                if item_name in modifiers and modifier in modifiers[item_name]:
                    menus[modifiers[item_name][modifier]] += row['Item Sold'] - row['Item Void']
                else:
                    menus[item_name] += row['Item Sold'] - row['Item Void']

                if row['Category'] == 'Paket Bukber':
                    if 'Ala-ala' in row['Item Name'] or 'Jomblo' in row['Item Name']:
                        menus['Iced Tea'] += (1 * (row['Item Sold'] - row['Item Void']))
                    elif 'Bucin' in row['Item Name']:
                        menus['Iced Tea'] += (2 * (row['Item Sold'] - row['Item Void']))
                    elif 'Gosip' in row['Item Name']:
                        menus['Iced Tea'] += (4 * (row['Item Sold'] - row['Item Void']))

                    if not pd.isna(row['Modifier']):
                        for bukber_menu_modifier in row['Modifier'].split(','):
                            bukber_menu, qty = bukber_menu_modifier.rsplit(' ', maxsplit=1)
                            menus[bukber_menu.rstrip('.')] += (int(qty.rstrip('X')) * (row['Item Sold'] - row['Item Void']))

        del menus['TOTAL']
        ws.cell(row=cell_row + 1, column=shifting_start_time.day + 1, value=sum(menus.values()))
        ws.cell(row=cell_row + 1, column=1, value=current_time.hour)
        print(menus)
        current_time = current_end_time + relativedelta(seconds=1)
        cell_row += 1

    current_time = shifting_start_time + relativedelta(days=1, hour=10)
    wb.save(filename)
